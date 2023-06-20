import pprint
import requests
from web3 import Web3
from config_networks import coins
from tqdm import tqdm
from tabulate import tabulate
import colorama
import openpyxl

colorama.init(autoreset=True)

### CONFIG ###
view_more_than = 0  # показать балансы монет где больше чем это число (usdt)
round_numbers = 4  # сколько чисел после запятой выводить
check_native_token = True  # показывать нативный токены сети True - да / False - нет
output_zero_balance = True  # выводить пустые баланы (True - да, False - нет)
output_in_excel = True
##############


def collect_coin(symbol, ampunt, amount_usdt):
    return {symbol: {"amount": ampunt, "usdt": amount_usdt}}


def collect_all_coins(balances):
    coin_totals = dict()
    for i in balances.values():
        for network, coins in i[0].items():
            if network != "GOERLI":
                for coin, details in coins.items():
                    if coin not in coin_totals:
                        coin_totals[coin] = {
                            "amount": details["amount"],
                            "usdt": details["usdt"],
                        }
                    else:
                        coin_totals[coin]["amount"] += details["amount"]
                        coin_totals[coin]["usdt"] += details["usdt"]
    return coin_totals


def collect_balance_one_address():
    balance = list()
    balance.append(dict())
    total_on_wallet = 0
    for network in coins:
        for key, value in network.items():
            w3 = Web3(Web3.HTTPProvider(value["rpc"]))
            balance[0][key] = dict()
            if check_native_token:
                amount = w3.eth.get_balance(i) / 10**18
                price_usdt = float(get_price_token(value["symbol"])) * amount
                total_on_wallet += price_usdt
                balance[0][key].update(
                    collect_coin(value["symbol"], amount, price_usdt)
                )
            for coin in value["coins"]:
                token = w3.eth.contract(
                    address=w3.to_checksum_address(coin["address"]), abi=coin["abi"]
                )
                amount = (
                    token.functions.balanceOf(w3.to_checksum_address(i)).call()
                    / 10 ** coin["decimal"]
                )
                price_usdt = float(get_price_token(coin["symbol"])) * amount
                total_on_wallet += price_usdt
                balance[0][key].update(collect_coin(coin["symbol"], amount, price_usdt))
        if "GOERLI" in balance[0]:
            total_on_wallet = total_on_wallet - balance[0]["GOERLI"]["ETH"]["usdt"]
        else:
            total_on_wallet = total_on_wallet
        balance.append({"total": total_on_wallet})
        return balance


def get_price_token(symbol):
    for i in prices_all_tokens:
        if i["currency_pair"] == f"{symbol}_USD":
            return i["last"]
        if i["currency_pair"] == f"{symbol}_USDT":
            return i["last"]
        if i["currency_pair"] == f"{symbol}_USDC":
            return i["last"]
    return 0


def output_balances(balances, filename=""):
    count = 1
    print("=" * 15)
    for wallet, networks in balances.items():
        if round(networks[1]["total"], round_numbers) == 0:
            continue
        print(f"{count}) {wallet}")
        for network, coins in networks[0].items():
            print(network)
            for key, value in coins.items():
                if not output_zero_balance and value["amount"] == 0.0:
                    continue
                if value["usdt"] >= view_more_than:
                    print(
                        f"\t{key.ljust(6)}{str(round(value['amount'],round_numbers))} | {round(value['usdt'],round_numbers)}$"
                    )
        print(f"На кошельке {round(networks[1]['total'],round_numbers)}$")
        count += 1
        print("=" * 15)


def tmp(balances):
    network_coins = {}

    for wallet, networks in balances.items():
        for network, coins in networks[0].items():
            if network not in network_coins:
                network_coins[network] = []
            network_coins[network].extend(coins.keys())
    for network in network_coins:
        network_coins[network] = list(set(network_coins[network]))

    return network_coins


def output_in_excel(balances):
    pprint.pprint(balances)
    book = openpyxl.Workbook()
    sheet = book.active
    netw_coin = tmp(balances)
    print(netw_coin)
    start_row = 1
    start_colm = 2
    for network, coin in netw_coin.items():
        sheet.cell(start_row, start_colm).value = network
        sheet.merge_cells(
            start_row=start_row,
            start_column=start_colm,
            end_row=start_row,
            end_column=start_colm + len(coin) - 1,
        )
        sheet
        start_colm += len(coin)

    start_row = 2
    start_colm = 2
    for network, coins in netw_coin.items():
        for coin in coins:
            sheet.cell(start_row, start_colm).value = coin
            start_colm += 1

    start_row = 3
    start_colm = 1
    for wallet, networks in balances.items():
        tuple_balance = [wallet]
        for network, coins in networks[0].items():
            for key, value in coins.items():
                tuple_balance.append(value["amount"])
        for i, value in enumerate(tuple_balance, start=1):
            
            sheet.cell(row=start_row, column=i).value = value
        start_row += 1
    book.save("test.xlsx")


def output_total_coins(coin_totals):
    for key, value in coin_totals.items():
        print(
            f"{key.ljust(6)}{round(value['amount'],round_numbers)} {round(value['usdt'],round_numbers)}$"
        )


def round_values(data):
    rounded_data = [
        [item[0], item[1], item[2], round(item[3], 1), round(item[4], 1)]
        for item in data
    ]
    return rounded_data


if __name__ == "__main__":
    prices_all_tokens = requests.get("https://api.gateio.ws/api/v4/spot/tickers").json()
    with open("./public.txt", "r") as wallets_txt:
        public_keys = wallets_txt.read().splitlines(keepends=False)
        public_keys = list(filter(None, public_keys))
    total_on_wallets = 0
    balances = dict()
    for i in tqdm(public_keys, desc="Fetching balances", unit="wallet"):
        balance_of_wallet = collect_balance_one_address()
        total_on_wallets += balance_of_wallet[1]["total"]
        balances[i] = balance_of_wallet
    output_in_excel(balances)
    # output_balances(balances)
    # print(f"Всего на кошельках {round(total_on_wallets,round_numbers)}$")
    # output_total_coins(collect_all_coins(balances))
